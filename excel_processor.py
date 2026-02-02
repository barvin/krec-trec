import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
from pathlib import Path
import os


def process_excel_file(input_file_path, output_file_path=None):
    """
    Process the Excel file according to specifications:
    1. Find the row with "Well Position" header
    2. Remove all rows above it
    3. Remove specific columns: Target Color, CQCONF, EXPFAIL, NOAMP
    4. Add columns: RNP/2, Copies/mln, Delta
    5. Calculate values for each sample
    """
    
    # Read the Excel file
    df = pd.read_excel(input_file_path, sheet_name=0, header=None)
    
    # Find the row containing "Well Position"
    header_row_idx = None
    for idx, row in df.iterrows():
        if any(cell == "Well Position" for cell in row if pd.notna(cell)):
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        raise ValueError("Could not find 'Well Position' header in the file")
    
    # Read the file again with the correct header
    df = pd.read_excel(input_file_path, sheet_name=0, skiprows=header_row_idx)
    
    # Remove unwanted columns
    columns_to_remove = ["Target Color", "CQCONF", "EXPFAIL", "NOAMP"]
    df = df.drop(columns=[col for col in columns_to_remove if col in df.columns], errors='ignore')
    
    # Add new columns
    df["RNP/2"] = None
    df["Copies/mln"] = None
    df["Delta"] = None
    
    # Process each sample
    # Group by Sample Name
    sample_groups = df.groupby("Sample Name", sort=False)
    
    for sample_name, group in sample_groups:
        # Get indices for this sample
        sample_indices = group.index
        
        # Find RNP row for this sample
        rnp_rows = group[group["Target Name"] == "RNP"]
        if not rnp_rows.empty:
            rnp_idx = rnp_rows.index[0]
            rnp_quantity = df.loc[rnp_idx, "Quantity"]
            rnp_ct = df.loc[rnp_idx, "CT"]
            
            if pd.notna(rnp_quantity):
                # Calculate RNP/2
                rnp_div_2 = rnp_quantity / 2
                df.loc[rnp_idx, "RNP/2"] = rnp_div_2
                                
                # Calculate Copies/mln for KREC
                krec_rows = group[group["Target Name"] == "KREC"]
                if not krec_rows.empty:
                    krec_idx = krec_rows.index[0]
                    krec_quantity = df.loc[krec_idx, "Quantity"]
                    if pd.notna(krec_quantity):
                        copies_mln_krec = krec_quantity / rnp_div_2 * 1_000_000
                        df.loc[krec_idx, "Copies/mln"] = copies_mln_krec
                
                # Calculate Copies/mln for TREC
                trec_rows = group[group["Target Name"] == "TREC"]
                if not trec_rows.empty:
                    trec_idx = trec_rows.index[0]
                    trec_quantity = df.loc[trec_idx, "Quantity"]
                    if pd.notna(trec_quantity):
                        copies_mln_trec = trec_quantity / rnp_div_2 * 1_000_000
                        df.loc[trec_idx, "Copies/mln"] = copies_mln_trec
            
            # Calculate Delta for KREC (KREC CT - RNP CT)
            krec_rows = group[group["Target Name"] == "KREC"]
            if not krec_rows.empty and pd.notna(rnp_ct):
                krec_idx = krec_rows.index[0]
                krec_ct = df.loc[krec_idx, "CT"]
                if pd.notna(krec_ct) and krec_ct != "Undetermined":
                    try:
                        delta_krec = float(krec_ct) - float(rnp_ct)
                        df.loc[krec_idx, "Delta"] = delta_krec
                    except (ValueError, TypeError):
                        pass
            
            # Calculate Delta for TREC (TREC CT - RNP CT)
            trec_rows = group[group["Target Name"] == "TREC"]
            if not trec_rows.empty and pd.notna(rnp_ct):
                trec_idx = trec_rows.index[0]
                trec_ct = df.loc[trec_idx, "CT"]
                if pd.notna(trec_ct) and trec_ct != "Undetermined":
                    try:
                        delta_trec = float(trec_ct) - float(rnp_ct)
                        df.loc[trec_idx, "Delta"] = delta_trec
                    except (ValueError, TypeError):
                        pass
    
    # Save the processed file
    if output_file_path is None:
        input_path = Path(input_file_path)
        output_file_path = input_path.parent / f"{input_path.stem}_processed{input_path.suffix}"
    
    df.to_excel(output_file_path, index=False, sheet_name="Results")
    
    # Apply borders to visually separate samples (every 4 rows)
    _apply_sample_borders(output_file_path, df)
    
    # Apply conditional formatting based on parameters
    _apply_conditional_formatting(output_file_path, df)
    
    return str(output_file_path)


def _apply_sample_borders(file_path, df):
    """
    Apply bold borders to separate each sample (every 4 rows).
    Each sample has 4 rows: KREC, RNP, SMN1, TREC.
    Also auto-sizes all columns to fit content.
    """
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Results']
    
    # Auto-size all columns
    for column in ws.columns:
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = 15
    
    # Create a thick border style for the bottom of each sample group
    thick_border = Border(
        bottom=Side(style='medium', color='000000')
    )
    
    # Track sample changes to apply borders
    current_sample = None
    sample_start_row = None
    row_count = 0
    
    # Iterate through data rows (skip header row 1)
    for idx, row_data in df.iterrows():
        sample_name = row_data['Sample Name']
        excel_row = idx + 2  # +2 because Excel is 1-indexed and we have a header row
        
        if current_sample != sample_name:
            # New sample detected
            if current_sample is not None and row_count > 0:
                # Apply border to the last row of the previous sample
                prev_row = excel_row - 1
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=prev_row, column=col)
                    cell.border = thick_border
            
            current_sample = sample_name
            sample_start_row = excel_row
            row_count = 1
        else:
            row_count += 1
    
    # Apply border to the last sample's last row
    if row_count > 0:
        last_row = len(df) + 1  # +1 for header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.border = thick_border
    
    # Save the workbook
    wb.save(file_path)
    wb.close()


def _load_parameters():
    """
    Load parameters from parameters.xlsx file.
    Returns a dictionary with threshold values.
    """
    params_file = 'parameters.xlsx'
    
    # Default values if file doesn't exist
    defaults = {
        'min_krec_copies': 10000.0,
        'min_trec_copies': 10000.0,
        'max_krec_delta': 11.5,
        'max_trec_delta': 12.0
    }
    
    if not os.path.exists(params_file):
        return defaults
    
    try:
        params_df = pd.read_excel(params_file)
        params = {}
        for _, row in params_df.iterrows():
            param_name = row['Parameter']
            param_value = row['Value']
            params[param_name] = param_value
        return params
    except Exception as e:
        print(f"Warning: Could not read parameters.xlsx: {e}. Using defaults.")
        return defaults


def _apply_conditional_formatting(file_path, df):
    """
    Apply conditional formatting based on parameters:
    - Yellow highlight for KREC Copies/mln if < min_krec_copies
    - Yellow highlight for TREC Copies/mln if < min_trec_copies
    - Yellow highlight for KREC Delta if > max_krec_delta
    - Yellow highlight for TREC Delta if > max_trec_delta
    """
    # Load parameters
    params = _load_parameters()
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Results']
    
    # Yellow fill for highlighting
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Get column indices
    header_row = [cell.value for cell in ws[1]]
    try:
        copies_col = header_row.index('Copies/mln') + 1
        delta_col = header_row.index('Delta') + 1
        target_col = header_row.index('Target Name') + 1
    except ValueError as e:
        print(f"Warning: Could not find required columns: {e}")
        wb.close()
        return
    
    # Iterate through data rows
    for idx, row_data in df.iterrows():
        excel_row = idx + 2  # +2 for Excel indexing and header
        target_name = row_data['Target Name']
        
        # Check KREC Copies/mln
        if target_name == 'KREC':
            copies_value = row_data.get('Copies/mln')
            if pd.notna(copies_value) and copies_value < params['min_krec_copies']:
                cell = ws.cell(row=excel_row, column=copies_col)
                cell.fill = yellow_fill
            
            # Check KREC Delta
            delta_value = row_data.get('Delta')
            if pd.notna(delta_value) and delta_value > params['max_krec_delta']:
                cell = ws.cell(row=excel_row, column=delta_col)
                cell.fill = yellow_fill
        
        # Check TREC Copies/mln
        elif target_name == 'TREC':
            copies_value = row_data.get('Copies/mln')
            if pd.notna(copies_value) and copies_value < params['min_trec_copies']:
                cell = ws.cell(row=excel_row, column=copies_col)
                cell.fill = yellow_fill
            
            # Check TREC Delta
            delta_value = row_data.get('Delta')
            if pd.notna(delta_value) and delta_value > params['max_trec_delta']:
                cell = ws.cell(row=excel_row, column=delta_col)
                cell.fill = yellow_fill
    
    # Save the workbook
    wb.save(file_path)
    wb.close()
